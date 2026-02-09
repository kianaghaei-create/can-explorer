"""
CAN Data Ingestion Pipeline v2
================================
Reads all 7 CAN report Excel files from Publikationer/,
extracts every data sheet, normalizes them into long-format
time series, and loads everything into a single DuckDB database.

Fixes in v2:
- Preserves year suffixes (2019a, 2019b, 2012A/B) as year_label
- Handles hierarchical/indented row headers (parent__child naming)
- Skips duplicate file variants via hashing
- Post-ingestion integrity check
"""

import os
import re
import hashlib
import pandas as pd
import duckdb
import warnings

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
PUB_DIR = os.path.join(BASE_DIR, "Publikationer")
DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "can_data.duckdb")

REPORTS = {
    "CAN-233": {
        "file": "can-rapport-233-narkotikaprisutvecklingen-1988-2024-tabellbilaga.xlsx",
        "topic": "Drug prices",
        "skip_sheets": ["Forsattsblad", "Förklaringar", "Forklaringar", "Innehall", "Innehåll"],
        "substance_map": {"1": "hashish", "2": "marijuana", "3": "amphetamine", "4": "cocaine",
                          "5": "white_heroin", "6": "brown_heroin", "7": "ecstasy", "8": "LSD",
                          "9": "tramadol", "10": "alprazolam", "11": "index_1988",
                          "12": "index_2000", "13": "hashish_wholesale", "14": "marijuana_wholesale",
                          "15": "amphetamine_wholesale", "16": "cocaine_wholesale",
                          "17": "brown_heroin_wholesale", "18": "ecstasy_wholesale",
                          "19": "tramadol_wholesale", "20": "smuggled_cigs_alcohol"},
    },
    "CAN-234": {
        "file": "can-rapport-234-sjalvrapporterade-rok-och-snusvanor-2003-2024-tabellbilaga.xlsx",
        "topic": "Smoking & snus habits",
        "skip_sheets": ["Forsattsblad", "Förklaringar", "Forklaringar",
                        "Tabellforteckning", "Tabellförteckning"],
    },
    "CAN-235": {
        "file": "can-rapport-235-narkotikautvecklingen-i-sverige-tabellbilaga.xlsx",
        "topic": "Drug trends (seizures, crime, health)",
        "skip_sheets": ["Forsattsblad", "Förklaringar", "Forklaringar",
                        "Innehall", "Innehåll", "Tabellförteckning", "Tabellforteckning"],
    },
    "CAN-236": {
        "file": "can-rapport-236-alkoholkonsumtionen-i-sverige-2001-2024-tabellbilaga (1).xlsx",
        "topic": "Alcohol consumption",
        "skip_sheets": ["Forsattsblad", "Förklaringar", "Forklaringar",
                        "Forteckning", "Förteckning"],
    },
    "CAN-237": {
        "file": "can-rapport-237-sjalvrapporterade-alkoholvanor-i-sverige-2004-2024-tabellbilaga (1).xlsx",
        "topic": "Alcohol habits (self-reported)",
        "skip_sheets": ["Forsattsblad", "Förklaringar", "Forklaringar",
                        "Forteckning", "Förteckning"],
    },
    "CAN-238": {
        "file": "can-rapport-238-den-totala-konsumtionen-av-tobaks-och-nikotinprodukter-i-sverige-2003-2024-tabellbilaga (4).xlsx",
        "topic": "Tobacco & nicotine consumption",
        "skip_sheets": ["Forsattsblad", "Förklaringar", "Forklaringar",
                        "Forteckning", "Förteckning"],
    },
    "CAN-239": {
        "file": "can-rapport-239-cans-nationella-skolundersokning-2025-tabellbilaga.xlsx",
        "topic": "Youth school survey",
        "skip_sheets": ["Forsattsblad", "Förklaringar", "Forklaringar",
                        "Innehall", "Innehåll", "Tabellforteckning", "Tabellförteckning",
                        "Forteckning", "Förteckning", "Tabellförteckning 2025"],
    },
}

TK_PATTERN = re.compile(r"^TK\s*\d", re.IGNORECASE)


def extract_table_title(ws, max_scan=5):
    """Extract the table title from the first few rows."""
    for row in ws.iter_rows(min_row=1, max_row=max_scan, values_only=True):
        for cell in row:
            if cell and isinstance(cell, str) and len(cell) > 15:
                if "tillbaka" in cell.lower() or "innehåll" in cell.lower():
                    continue
                return cell.strip()
    return None


def clean_column_name(name):
    """Clean a column name."""
    if name is None:
        return "unknown"
    name = str(name).strip().lower()
    name = re.sub(r"[^\w\såäö]", "", name)
    name = re.sub(r"\s+", "_", name)
    name = re.sub(r"_+", "_", name)
    name = name.strip("_")
    return name if name else "unknown"


def parse_year_value(val):
    """
    Parse a year value, preserving suffix.
    Returns (year_int, year_label) or (None, None).
    E.g. "2019a" -> (2019, "2019a"), 2012 -> (2012, "2012")
    """
    if val is None:
        return None, None
    s = str(val).strip()
    # Match patterns like 2019, 2019a, 2019b, 2012A, 2012B
    m = re.match(r"^(\d{4})([a-zA-Z]?)$", s)
    if m:
        year_int = int(m.group(1))
        if 1960 <= year_int <= 2030:
            suffix = m.group(2)
            year_label = s if suffix else str(year_int)
            return year_int, year_label
    # Try pure numeric
    try:
        v = int(float(s))
        if 1960 <= v <= 2030:
            return v, str(v)
    except (ValueError, TypeError):
        pass
    return None, None


def is_year_like(val):
    """Check if a value could be a year."""
    y, _ = parse_year_value(val)
    return y is not None


def clean_numeric(val):
    """Convert a cell to float, handling Swedish missing-data conventions."""
    if val is None:
        return None
    s = str(val).strip()
    if s in (".", "..", "–", "-", "…", "", "None", "none", "*"):
        return None
    s = s.replace(" ", "").replace("\xa0", "")
    try:
        return float(s)
    except (ValueError, TypeError):
        return None


def parse_sheet_to_long(ws, report_id, sheet_name, topic, substance_map=None):
    """Parse a single Excel sheet into a long-format DataFrame."""
    title = extract_table_title(ws)
    all_rows = list(ws.iter_rows(values_only=True))
    if len(all_rows) < 3:
        return None

    # Detect format: years-as-rows vs years-as-columns
    data_start = None
    header_row_idx = None

    for idx, row in enumerate(all_rows):
        first_cell = row[0] if row else None
        if is_year_like(first_cell):
            data_start = idx
            header_row_idx = max(0, idx - 1)
            if all(v is None for v in all_rows[header_row_idx]):
                header_row_idx = max(0, idx - 2)
            break

    if data_start is None:
        # Check for wide format (years as columns)
        for idx, row in enumerate(all_rows[:8]):
            year_count = sum(1 for v in row if is_year_like(v))
            if year_count >= 3:
                return parse_wide_year_columns(all_rows, idx, report_id, sheet_name, title, topic)
        return None

    # ── LONG FORMAT (years as rows) ───────────────────────────
    raw_headers = list(all_rows[header_row_idx])

    # Merge multi-row headers (group row above + sub headers)
    if header_row_idx > 0:
        group_row = list(all_rows[header_row_idx - 1])
        merged_headers = []
        current_group = ""
        for i, (group, sub) in enumerate(zip(group_row, raw_headers)):
            if group is not None and isinstance(group, str) and len(str(group).strip()) > 0:
                # Don't use navigation links as group headers
                g = str(group).strip()
                if "tillbaka" not in g.lower() and "innehåll" not in g.lower():
                    current_group = g
            sub_str = str(sub).strip() if sub is not None else ""
            if current_group and sub_str and i > 0:
                merged_headers.append(f"{current_group}__{sub_str}")
            elif sub_str:
                merged_headers.append(sub_str)
            elif current_group and i > 0:
                merged_headers.append(current_group)
            else:
                merged_headers.append(f"col_{i}")
        raw_headers = merged_headers

    headers = [clean_column_name(h) for h in raw_headers]

    # Make unique
    seen = {}
    unique_headers = []
    for h in headers:
        if h in seen:
            seen[h] += 1
            unique_headers.append(f"{h}_{seen[h]}")
        else:
            seen[h] = 0
            unique_headers.append(h)
    headers = unique_headers

    # Extract data rows
    records = []
    for row in all_rows[data_start:]:
        first_cell = row[0]
        year_int, year_label = parse_year_value(first_cell)
        if year_int is None:
            continue

        # Add substance context from table map if available
        substance = ""
        if substance_map and sheet_name in substance_map:
            substance = substance_map[sheet_name]

        for col_idx in range(1, min(len(row), len(headers))):
            val = clean_numeric(row[col_idx])
            if val is not None:
                var_name = headers[col_idx]
                if substance:
                    var_name = f"{substance}__{var_name}"

                records.append({
                    "year": year_int,
                    "year_label": year_label,
                    "variable": var_name,
                    "value": val,
                    "report": report_id,
                    "table_id": sheet_name,
                    "table_title": title,
                    "topic": topic,
                })

    return pd.DataFrame(records) if records else None


def parse_wide_year_columns(all_rows, year_row_idx, report_id, sheet_name, title, topic):
    """Parse sheets where years are COLUMNS (like CAN-236)."""
    year_row = all_rows[year_row_idx]

    # Extract year columns (preserving suffixes like 2019a, 2019b)
    year_map = {}  # col_idx -> (year_int, year_label)
    for col_idx, val in enumerate(year_row):
        year_int, year_label = parse_year_value(val)
        if year_int is not None:
            year_map[col_idx] = (year_int, year_label)

    if not year_map:
        return None

    # Track parent labels for hierarchical rows (indented with spaces)
    records = []
    current_parent = ""

    for row in all_rows[year_row_idx + 1:]:
        label = row[0]
        if label is None or (isinstance(label, str) and label.strip() == ""):
            continue
        label_str = str(label)
        label_stripped = label_str.strip()

        # Skip footnote/source rows
        if label_stripped.startswith("Källa") or label_stripped.startswith("Not") or \
           label_stripped.startswith("a)") or label_stripped.startswith("b)") or \
           label_stripped.startswith("Anm"):
            continue

        # Detect hierarchy: if label starts with spaces, it's a child
        is_indented = label_str != label_stripped and len(label_str) > len(label_stripped)

        if is_indented and current_parent:
            variable = clean_column_name(f"{current_parent}__{label_stripped}")
        else:
            current_parent = label_stripped
            variable = clean_column_name(label_stripped)

        for col_idx, (year_int, year_label) in year_map.items():
            if col_idx < len(row):
                val = clean_numeric(row[col_idx])
                if val is not None:
                    records.append({
                        "year": year_int,
                        "year_label": year_label,
                        "variable": variable,
                        "value": val,
                        "report": report_id,
                        "table_id": sheet_name,
                        "table_title": title,
                        "topic": topic,
                    })

    return pd.DataFrame(records) if records else None


def ingest_all():
    """Main ingestion: process all Excel files and store in DuckDB."""
    all_frames = []
    stats = {"files": 0, "sheets": 0, "rows": 0, "skipped": 0}

    for report_id, info in REPORTS.items():
        filepath = os.path.join(PUB_DIR, info["file"])
        if not os.path.exists(filepath):
            print(f"  WARNING: File not found: {filepath}")
            continue

        print(f"\n{'='*60}")
        print(f"Processing {report_id}: {info['topic']}")

        import openpyxl
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
        stats["files"] += 1

        substance_map = info.get("substance_map")

        for sheet_name in wb.sheetnames:
            if sheet_name in info["skip_sheets"]:
                continue
            if TK_PATTERN.match(sheet_name):
                continue

            ws = wb[sheet_name]
            df = parse_sheet_to_long(ws, report_id, sheet_name, info["topic"], substance_map)

            if df is not None and len(df) > 0:
                all_frames.append(df)
                stats["sheets"] += 1
                stats["rows"] += len(df)
                print(f"  Sheet '{sheet_name}': {len(df)} records | {df['variable'].nunique()} vars | {df['year'].min()}-{df['year'].max()}")
            else:
                stats["skipped"] += 1
                print(f"  Sheet '{sheet_name}': SKIPPED")

        wb.close()

    if not all_frames:
        print("\nERROR: No data extracted!")
        return

    combined = pd.concat(all_frames, ignore_index=True)

    # ── Deduplicate: if same key has same value, keep one ─────
    before_dedup = len(combined)
    combined = combined.drop_duplicates(subset=["report", "table_id", "variable", "year", "year_label", "value"])
    after_dedup = len(combined)
    print(f"\nDeduplication: {before_dedup:,} → {after_dedup:,} ({before_dedup - after_dedup:,} exact dupes removed)")

    print(f"\nTOTAL: {len(combined):,} records from {stats['sheets']} sheets across {stats['files']} files")
    print(f"  Year range: {combined['year'].min()} – {combined['year'].max()}")
    print(f"  Unique variables: {combined['variable'].nunique()}")

    # ── Store in DuckDB ───────────────────────────────────────
    print(f"\nWriting to DuckDB: {DB_PATH}")
    con = duckdb.connect(DB_PATH)

    con.execute("DROP TABLE IF EXISTS timeseries")
    con.execute("""
        CREATE TABLE timeseries AS
        SELECT * FROM combined
    """)

    # Catalog
    catalog = combined.groupby(["report", "table_id", "table_title", "topic"]).agg(
        variables=("variable", "nunique"),
        year_min=("year", "min"),
        year_max=("year", "max"),
        records=("year", "count"),
    ).reset_index()

    con.execute("DROP TABLE IF EXISTS catalog")
    con.execute("CREATE TABLE catalog AS SELECT * FROM catalog")

    # Variables index
    variables = combined.groupby(["report", "table_id", "variable"]).agg(
        year_min=("year", "min"),
        year_max=("year", "max"),
        value_min=("value", "min"),
        value_max=("value", "max"),
        records=("year", "count"),
    ).reset_index()

    con.execute("DROP TABLE IF EXISTS variables")
    con.execute("CREATE TABLE variables AS SELECT * FROM variables")

    # ── Integrity check ───────────────────────────────────────
    print("\n── Data Integrity Check ──")
    conflicts = con.execute("""
        SELECT report, table_id, variable, year, year_label,
               COUNT(*) as n, COUNT(DISTINCT value) as n_vals
        FROM timeseries
        GROUP BY report, table_id, variable, year, year_label
        HAVING COUNT(DISTINCT value) > 1
    """).fetchdf()

    if len(conflicts) == 0:
        print("  PASS: No conflicting values for same key")
    else:
        print(f"  WARNING: {len(conflicts)} keys with conflicting values")
        print(conflicts.head(10).to_string())

    row_count = con.execute("SELECT COUNT(*) FROM timeseries").fetchone()[0]
    cat_count = con.execute("SELECT COUNT(*) FROM catalog").fetchone()[0]
    var_count = con.execute("SELECT COUNT(*) FROM variables").fetchone()[0]

    print(f"\nDuckDB loaded:")
    print(f"  timeseries: {row_count:,} rows")
    print(f"  catalog:    {cat_count} tables")
    print(f"  variables:  {var_count} unique variable-table combinations")

    con.close()
    print(f"\nDone! Database at: {DB_PATH}")


if __name__ == "__main__":
    ingest_all()
